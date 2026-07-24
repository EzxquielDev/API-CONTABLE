import io
from datetime import date
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from auth import require_api_key
from services.kardex_service import obtener_kardex

kardex_bp = Blueprint("kardex", __name__, url_prefix="/api")


def _validar_fechas():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    if not desde or not hasta:
        return None, None, ("Faltan parámetros 'desde' y 'hasta' (formato YYYY-MM-DD)", 400)
    try:
        date.fromisoformat(desde)
        date.fromisoformat(hasta)
    except ValueError:
        return None, None, ("Formato de fecha inválido, usa YYYY-MM-DD", 400)
    return desde, hasta, None


@kardex_bp.route("/kardex", methods=["GET"])
@require_api_key
def kardex():
    desde, hasta, error = _validar_fechas()
    if error:
        return jsonify({"error": error[0]}), error[1]
    codigo = request.args.get("codigo")
    try:
        data = obtener_kardex(desde, hasta, codigo_producto=codigo)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@kardex_bp.route("/kardex.xlsx", methods=["GET"])
@require_api_key
def kardex_xlsx():
    desde, hasta, error = _validar_fechas()
    if error:
        return jsonify({"error": error[0]}), error[1]
    codigo = request.args.get("codigo")

    try:
        data = obtener_kardex(desde, hasta, codigo_producto=codigo)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "KARDEX"

    font_titulo = Font(name="Arial", size=12, bold=True)
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    font_normal = Font(name="Arial", size=10)
    font_saldo_inicial = Font(name="Arial", size=10, italic=True)

    ws["A1"] = "REPORTE DE MOVIMIENTOS DE INVENTARIO (KARDEX DE ARTICULOS)"
    ws["A1"].font = font_titulo
    ws.merge_cells("A1:K1")

    encabezados = ["FECHA", "CODIGO", "DESCRIPCION", "Qty entrada", "costo unitario",
                    "costo total", "Qty salida", "costo de venta (total)", "existencia", "saldo", "referencia"]
    fila = 3
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=col, value=texto)
        celda.font = font_header
        celda.fill = fill_header
        celda.alignment = Alignment(horizontal="center")
    fila += 1

    for producto in data["productos"]:
        # Fila de saldo inicial (igual que en tu archivo de referencia)
        ws.cell(row=fila, column=2, value=producto["codigo"]).font = font_saldo_inicial
        ws.cell(row=fila, column=3, value=producto["descripcion"]).font = font_saldo_inicial
        ws.cell(row=fila, column=9, value=producto["existencia_inicial"]).font = font_saldo_inicial
        ws.cell(row=fila, column=10, value=producto["saldo_inicial"]).font = font_saldo_inicial
        ws.cell(row=fila, column=11, value="saldo inicial").font = font_saldo_inicial
        fila += 1

        for m in producto["movimientos"]:
            ws.cell(row=fila, column=1, value=m["fecha"]).font = font_normal
            ws.cell(row=fila, column=2, value=producto["codigo"]).font = font_normal
            ws.cell(row=fila, column=3, value=producto["descripcion"]).font = font_normal
            ws.cell(row=fila, column=4, value=m["qty_entrada"] or None).font = font_normal
            ws.cell(row=fila, column=5, value=m["costo_unitario"]).font = font_normal
            ws.cell(row=fila, column=6, value=m["costo_total"]).font = font_normal
            ws.cell(row=fila, column=7, value=m["qty_salida"] or None).font = font_normal
            ws.cell(row=fila, column=8, value=m["precio_venta"]).font = (
                Font(name="Arial", size=10, italic=True, color="9C6500") if m.get("precio_estimado") else font_normal
            )
            ws.cell(row=fila, column=9, value=m["existencia"]).font = font_normal
            ws.cell(row=fila, column=10, value=m["saldo"]).font = font_normal
            ws.cell(row=fila, column=11, value=m["referencia"]).font = font_normal
            for col in (5, 6, 8, 10):
                ws.cell(row=fila, column=col).number_format = "$#,##0.00"
            fila += 1

        fila += 1  # línea en blanco entre productos, igual que el original

    anchos = {"A": 12, "B": 14, "C": 45, "D": 12, "E": 14, "F": 12, "G": 10,
               "H": 14, "I": 11, "J": 11, "K": 16}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"kardex_{data['desde']}_a_{data['hasta']}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo,
    )
