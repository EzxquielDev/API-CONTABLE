import csv
import io
from flask import Blueprint, jsonify, Response, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from auth import require_api_key
from services.dashboard_service import (
    obtener_resumen_financiero,
    obtener_facturas_pendientes,
    obtener_estado_financiero,
)

dashboard_bp = Blueprint("dashboard", __name__, url_prefix="/api/dashboard")


@dashboard_bp.route("/resumen", methods=["GET"])
@require_api_key
def resumen():
    """Resumen financiero del mes actual. Ideal para Google Sheets (Apps Script) o Power Query."""
    try:
        data = obtener_resumen_financiero()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/facturas-pendientes", methods=["GET"])
@require_api_key
def facturas_pendientes():
    """Facturas de cliente con saldo pendiente, en JSON."""
    try:
        data = obtener_facturas_pendientes(tipo="cliente")
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/facturas-pendientes.csv", methods=["GET"])
@require_api_key
def facturas_pendientes_csv():
    """Mismo contenido que arriba, pero en CSV listo para abrir/pegar en Excel."""
    try:
        data = obtener_facturas_pendientes(tipo="cliente")

        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=facturas_pendientes.csv"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/estado-financiero", methods=["GET"])
@require_api_key
def estado_financiero_json():
    """Estado financiero (ventas vs. compras) en JSON, para mostrarlo en el frontend."""
    try:
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        data = obtener_estado_financiero(desde=desde, hasta=hasta)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/estado-financiero.xlsx", methods=["GET"])
@require_api_key
def estado_financiero_xlsx():
    """Estado financiero (ventas vs. compras) en Excel para un rango de fechas.

    Parámetros opcionales: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    Si no se envían, se usa el mes actual.
    """
    try:
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        data = obtener_estado_financiero(desde=desde, hasta=hasta)

        wb = Workbook()
        ws = wb.active
        ws.title = "Estado Financiero"

        negrita = Font(bold=True)
        fuente_encabezado = Font(bold=True, color="FFFFFF")
        relleno_encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        formato_moneda = "#,##0.00"

        ws["A1"] = "Estado Financiero"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        ws["A2"] = "Período"
        ws["B2"] = f'{data["desde"]} a {data["hasta"]}'
        ws["A2"].font = negrita

        def escribir_seccion(fila, titulo, filas_datos):
            ws.cell(row=fila, column=1, value=titulo).font = fuente_encabezado
            ws.cell(row=fila, column=1).fill = relleno_encabezado
            ws.cell(row=fila, column=2).fill = relleno_encabezado
            fila += 1
            for etiqueta, valor in filas_datos:
                ws.cell(row=fila, column=1, value=etiqueta)
                celda = ws.cell(row=fila, column=2, value=valor)
                if isinstance(valor, float):
                    celda.number_format = formato_moneda
                fila += 1
            return fila + 1

        ventas = data["ventas"]
        compras = data["compras"]

        fila = 4
        fila = escribir_seccion(fila, "VENTAS", [
            ("Total facturado", ventas["total_facturado"]),
            ("Total cobrado", ventas["total_cobrado"]),
            ("Total por cobrar", ventas["total_por_cobrar"]),
            ("Cantidad de facturas", ventas["cantidad_facturas"]),
        ])

        fila = escribir_seccion(fila, "COMPRAS (Órdenes de Compra)", [
            ("Total ordenado", compras["total_ordenado"]),
            ("Ya facturado por proveedor", compras["total_facturado"]),
            ("Pendiente de facturar", compras["total_por_facturar"]),
            ("Cantidad de órdenes", compras["cantidad_ordenes"]),
        ])

        ws.cell(row=fila, column=1, value="RESULTADO NETO").font = negrita
        celda_resultado = ws.cell(row=fila, column=2, value=data["resultado_neto"])
        celda_resultado.font = negrita
        celda_resultado.number_format = formato_moneda

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'estado_financiero_{data["desde"]}_a_{data["hasta"]}.xlsx'
        return Response(
            buffer.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@dashboard_bp.route("/gastos-pendientes", methods=["GET"])
@require_api_key
def gastos_pendientes():
    """Facturas de proveedor con saldo pendiente, en JSON."""
    try:
        data = obtener_facturas_pendientes(tipo="proveedor")
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
