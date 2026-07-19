import csv
import io
from datetime import date, timedelta

from flask import Blueprint, Response, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from auth import require_api_key
from services.inventario_service import (
    obtener_almacenes,
    obtener_entradas_inventario,
    obtener_productos_inventario,
    obtener_reporte_inventario,
    obtener_resumen_inventario,
    obtener_todas_entradas_inventario,
)


inventario_bp = Blueprint("inventario", __name__, url_prefix="/api/inventario")


def _filtros():
    almacen_id = request.args.get("almacen_id", type=int)
    producto = request.args.get("producto", default="", type=str)
    return almacen_id, producto


def _reporte_completo(almacen_id, producto):
    return obtener_productos_inventario(almacen_id, producto)


def _fechas_entradas():
    """Obtiene y valida el periodo usado por las entradas de inventario."""
    desde = request.args.get("desde", default=(date.today() - timedelta(days=30)).isoformat())
    hasta = request.args.get("hasta", default=date.today().isoformat())
    date.fromisoformat(desde)
    date.fromisoformat(hasta)
    if desde > hasta:
        raise ValueError("La fecha inicial no puede ser posterior a la fecha final.")
    return desde, hasta


@inventario_bp.route("/almacenes", methods=["GET"])
@require_api_key
def almacenes():
    try:
        return jsonify({"almacenes": obtener_almacenes()})
    except Exception as error:
        return jsonify({"error": str(error)}), 500


@inventario_bp.route("/reporte", methods=["GET"])
@require_api_key
def reporte():
    almacen_id, producto = _filtros()
    pagina = request.args.get("pagina", default=1, type=int)
    por_pagina = request.args.get("por_pagina", default=100, type=int)
    try:
        return jsonify(obtener_reporte_inventario(almacen_id, producto, pagina, por_pagina))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": str(error)}), 500


@inventario_bp.route("/resumen", methods=["GET"])
@require_api_key
def resumen():
    try:
        almacen_id, producto = _filtros()
        return jsonify(obtener_resumen_inventario(almacen_id, producto))
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": str(error)}), 500


@inventario_bp.route("/entradas", methods=["GET"])
@require_api_key
def entradas():
    """Compras facturadas por proveedor y producto."""
    limite = request.args.get("limite", default=200, type=int)
    try:
        desde, hasta = _fechas_entradas()
        return jsonify({"desde": desde, "hasta": hasta, "entradas": obtener_entradas_inventario(desde, hasta, limite)})
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": str(error)}), 500


@inventario_bp.route("/entradas.xlsx", methods=["GET"])
@require_api_key
def entradas_xlsx():
    """Descarga las entradas de productos del periodo seleccionado en Excel."""
    try:
        desde, hasta = _fechas_entradas()
        entradas = obtener_todas_entradas_inventario(desde, hasta)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception as error:
        return jsonify({"error": str(error)}), 500

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Entradas"
    encabezados = ["Fecha", "Proveedor", "Factura", "Código", "Descripción", "Cantidad", "Unitario", "Total"]
    estilo_encabezado = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=texto)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = estilo_encabezado
        celda.alignment = Alignment(horizontal="center")

    for fila, entrada in enumerate(entradas, start=2):
        valores = [
            entrada.get("fecha"), entrada.get("proveedor"), entrada.get("factura"), entrada.get("codigo"),
            entrada.get("producto"), entrada.get("cantidad"), entrada.get("costo_unitario"), entrada.get("valor_total"),
        ]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        hoja.cell(row=fila, column=6).number_format = "#,##0.00"
        for columna in (7, 8):
            hoja.cell(row=fila, column=columna).number_format = "$#,##0.00"

    for columna, ancho in {"A": 14, "B": 22, "C": 18, "D": 18, "E": 42, "F": 14, "G": 16, "H": 16}.items():
        hoja.column_dimensions[columna].width = ancho
    hoja.freeze_panes = "A2"

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"entradas_{desde}_a_{hasta}.xlsx",
    )


@inventario_bp.route("/reporte.csv", methods=["GET"])
@require_api_key
def reporte_csv():
    try:
        productos = _reporte_completo(*_filtros())
        salida = io.StringIO()
        columnas = ["sku", "producto", "categoria", "unidad_medida", "existencia", "reservado", "disponible", "costo_unitario", "precio_venta", "valor_inventario"]
        writer = csv.DictWriter(salida, fieldnames=columnas)
        writer.writeheader()
        writer.writerows(productos)
        return Response(salida.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=inventario.csv"})
    except Exception as error:
        return jsonify({"error": str(error)}), 500


@inventario_bp.route("/reporte.xlsx", methods=["GET"])
@require_api_key
def reporte_xlsx():
    try:
        productos = _reporte_completo(*_filtros())
    except Exception as error:
        return jsonify({"error": str(error)}), 500

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Inventario"
    encabezados = ["SKU", "Producto", "Categoría", "Unidad", "Existencia", "Reservado", "Disponible", "Costo unitario", "Precio venta", "Valor inventario"]
    estilo_encabezado = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=texto)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = estilo_encabezado
        celda.alignment = Alignment(horizontal="center")

    for fila, producto in enumerate(productos, start=2):
        valores = [producto[campo] for campo in ["sku", "producto", "categoria", "unidad_medida", "existencia", "reservado", "disponible", "costo_unitario", "precio_venta", "valor_inventario"]]
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
        for columna in (8, 9, 10):
            hoja.cell(row=fila, column=columna).number_format = "$#,##0.00"

    for columna, ancho in {"A": 18, "B": 38, "C": 25, "D": 14, "E": 13, "F": 13, "G": 13, "H": 16, "I": 16, "J": 18}.items():
        hoja.column_dimensions[columna].width = ancho
    hoja.freeze_panes = "A2"

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="inventario.xlsx")
