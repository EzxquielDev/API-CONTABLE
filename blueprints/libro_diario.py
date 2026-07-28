import io
from flask import Blueprint, jsonify, Response, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from auth import require_api_key
from services.dashboard_service import obtener_libro_diario

libro_diario_bp = Blueprint("libro_diario", __name__, url_prefix="/api")


@libro_diario_bp.route("/libro-diario", methods=["GET"])
@require_api_key
def libro_diario_json():
    """Libro Diario en JSON: todos los apuntes contables (débito/crédito) del período."""
    try:
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        data = obtener_libro_diario(desde=desde, hasta=hasta)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@libro_diario_bp.route("/libro-diario.xlsx", methods=["GET"])
@require_api_key
def libro_diario_xlsx():
    """Libro Diario en Excel, con columnas Fecha / Asiento / Diario / Cuenta / Etiqueta / Debe / Haber."""
    try:
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        data = obtener_libro_diario(desde=desde, hasta=hasta)

        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"

        negrita = Font(bold=True)
        fuente_encabezado = Font(bold=True, color="FFFFFF")
        relleno_encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        formato_moneda = "#,##0.00"

        ws["A1"] = "Libro Diario"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:G1")

        ws["A2"] = "Período"
        ws["B2"] = f'{data["desde"]} a {data["hasta"]}'
        ws["A2"].font = negrita

        encabezados = ["Fecha", "Asiento", "Diario", "Cuenta", "Etiqueta", "Debe", "Haber"]
        fila_encabezado = 4
        for col, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
            celda.font = fuente_encabezado
            celda.fill = relleno_encabezado

        fila = fila_encabezado + 1
        for m in data["movimientos"]:
            ws.cell(row=fila, column=1, value=m["fecha"])
            ws.cell(row=fila, column=2, value=m["asiento"])
            ws.cell(row=fila, column=3, value=m["diario"])
            ws.cell(row=fila, column=4, value=m["cuenta"])
            ws.cell(row=fila, column=5, value=m["etiqueta"])
            c_debe = ws.cell(row=fila, column=6, value=m["debe"])
            c_debe.number_format = formato_moneda
            c_haber = ws.cell(row=fila, column=7, value=m["haber"])
            c_haber.number_format = formato_moneda
            fila += 1

        fila += 1
        ws.cell(row=fila, column=5, value="TOTALES").font = negrita
        c_total_debe = ws.cell(row=fila, column=6, value=data["total_debe"])
        c_total_debe.font = negrita
        c_total_debe.number_format = formato_moneda
        c_total_haber = ws.cell(row=fila, column=7, value=data["total_haber"])
        c_total_haber.font = negrita
        c_total_haber.number_format = formato_moneda

        anchos = {"A": 12, "B": 20, "C": 16, "D": 26, "E": 30, "F": 14, "G": 14}
        for columna, ancho in anchos.items():
            ws.column_dimensions[columna].width = ancho

        ws.freeze_panes = f"A{fila_encabezado + 1}"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nombre_archivo = f'libro_diario_{data["desde"]}_a_{data["hasta"]}.xlsx'
        return Response(
            buffer.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
