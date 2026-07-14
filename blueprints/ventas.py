import io
from datetime import date
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from auth import require_api_key
from services.ventas_service import obtener_reporte_ventas, obtener_productos_mas_vendidos

ventas_bp = Blueprint("ventas", __name__, url_prefix="/api/ventas")


def _validar_fechas():
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")
    if not desde or not hasta:
        return None, None, ("Faltan parámetros 'desde' y 'hasta' (formato YYYY-MM-DD)", 400)
    try:
        date.fromisoformat(desde)
        date.fromisoformat(hasta)
    except ValueError:
        return None, None, ("Formato de fecha inválido, usa YYYY-MM-DD", 400)
    return desde, hasta, None


@ventas_bp.route("/reporte", methods=["GET"])
@require_api_key
def reporte():
    desde, hasta, error = _validar_fechas()
    if error:
        return jsonify({"error": error[0]}), error[1]
    try:
        data = obtener_reporte_ventas(desde, hasta)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@ventas_bp.route("/productos-mas-vendidos", methods=["GET"])
@require_api_key
def productos_mas_vendidos():
    desde, hasta, error = _validar_fechas()
    if error:
        return jsonify({"error": error[0]}), error[1]
    limite = request.args.get("limite", default=20, type=int)
    try:
        data = obtener_productos_mas_vendidos(desde, hasta, limite=limite)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@ventas_bp.route("/reporte.xlsx", methods=["GET"])
@require_api_key
def reporte_xlsx():
    desde, hasta, error = _validar_fechas()
    if error:
        return jsonify({"error": error[0]}), error[1]

    try:
        data = obtener_reporte_ventas(desde, hasta)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    font_normal = Font(name="Arial", size=10)
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    font_total = Font(name="Arial", size=10, bold=True)

    # Título y rango
    ws["A1"] = f"Reporte de Ventas — {data['desde']} a {data['hasta']}"
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws.merge_cells("A1:F1")

    # Encabezados
    encabezados = ["Número", "Cliente", "Fecha", "Subtotal sin IVA", "IVA", "Total", "Estado de pago"]
    fila_encabezado = 3
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=texto)
        celda.font = font_header
        celda.fill = fill_header
        celda.alignment = Alignment(horizontal="center")

    # Detalle
    fila = fila_encabezado + 1
    primera_fila_datos = fila
    for f in data["facturas"]:
        ws.cell(row=fila, column=1, value=f["numero"]).font = font_normal
        ws.cell(row=fila, column=2, value=f["cliente"]).font = font_normal
        ws.cell(row=fila, column=3, value=f["fecha"]).font = font_normal
        ws.cell(row=fila, column=4, value=f["subtotal_sin_iva"]).font = font_normal
        ws.cell(row=fila, column=5, value=f["iva"]).font = font_normal
        ws.cell(row=fila, column=6, value=f["total"]).font = font_normal
        ws.cell(row=fila, column=7, value=f["estado_pago"]).font = font_normal
        for col in (4, 5, 6):
            ws.cell(row=fila, column=col).number_format = "$#,##0.00"
        fila += 1
    ultima_fila_datos = fila - 1

    # Totales (con fórmulas, para que se recalculen si se edita el detalle)
    fila_totales = fila + 1
    ws.cell(row=fila_totales, column=3, value="TOTALES").font = font_total
    if data["facturas"]:
        ws.cell(row=fila_totales, column=4,
                value=f"=SUM(D{primera_fila_datos}:D{ultima_fila_datos})").font = font_total
        ws.cell(row=fila_totales, column=5,
                value=f"=SUM(E{primera_fila_datos}:E{ultima_fila_datos})").font = font_total
        ws.cell(row=fila_totales, column=6,
                value=f"=SUM(F{primera_fila_datos}:F{ultima_fila_datos})").font = font_total
    else:
        ws.cell(row=fila_totales, column=4, value=0).font = font_total
        ws.cell(row=fila_totales, column=5, value=0).font = font_total
        ws.cell(row=fila_totales, column=6, value=0).font = font_total
    for col in (4, 5, 6):
        ws.cell(row=fila_totales, column=col).number_format = "$#,##0.00"

    # Ancho de columnas
    anchos = {"A": 16, "B": 30, "C": 14, "D": 18, "E": 14, "F": 14, "G": 16}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"reporte_ventas_{data['desde']}_a_{data['hasta']}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_archivo,
    )
